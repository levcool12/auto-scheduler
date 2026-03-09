import json
import random
import pandas as pd
from flask import Flask, request, send_file, render_template_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

app = Flask(__name__)

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri"]
SLOTS = [1, 2, 3, 4, 5]

HTML = """
<h2>Schedule Generator</h2>
<form method="post" enctype="multipart/form-data">
<input type="file" name="file" accept=".json">
<input type="submit" value="Generate Schedule">
</form>
"""

#чтение входного файла

def parse_json(file):

    data = json.load(file)

    groups = data["groups"]
    rooms = data["rooms"]
    classes = data["classes"]

    return groups, rooms, classes

#генерация расписания 

def generate_schedule(groups, rooms, classes):

    teacher_busy = {}
    room_busy = {}
    group_busy = {}

    schedule = []

    for cls in classes:

        group = cls["group"]
        subject = cls["subject"]
        teacher = cls["teacher"]
        hours = cls["hours"]

        for _ in range(hours):

            placed = False

            while not placed:

                day = random.choice(DAYS)
                slot = random.choice(SLOTS)
                room = random.choice(rooms)

                key = (day, slot)

                if (
                    teacher_busy.get((teacher, key)) or
                    room_busy.get((room, key)) or
                    group_busy.get((group, key))
                ):
                    continue

                teacher_busy[(teacher, key)] = True
                room_busy[(room, key)] = True
                group_busy[(group, key)] = True

                schedule.append({
                    "Group": group,
                    "Subject": subject,
                    "Teacher": teacher,
                    "Room": room,
                    "Day": day,
                    "Slot": slot
                })

                placed = True

    return pd.DataFrame(schedule)

#экспорт расписания в excel таблицу

def export_excel(schedule, groups):

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    header = ["Day", "Slot"] + groups
    ws.append(header)

    for col in ws[1]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal="center")

    for day in DAYS:

        start_row = ws.max_row + 1

        for slot in SLOTS:

            row = [day, slot]

            for group in groups:

                lesson = schedule[
                    (schedule["Day"] == day) &
                    (schedule["Slot"] == slot) &
                    (schedule["Group"] == group)
                ]

                if not lesson.empty:

                    l = lesson.iloc[0]

                    text = f"{l['Subject']}, {l['Room']}, {l['Teacher']}"

                else:
                    text = ""

                row.append(text)

            ws.append(row)

        end_row = ws.max_row

        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=end_row, end_column=1)

        ws.cell(row=start_row, column=1).alignment = Alignment(
            vertical="center",
            horizontal="center"
        )

    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = 20

    filename = "schedule.xlsx"
    wb.save(filename)

    return filename

#интерфейс веб приложения

@app.route("/", methods=["GET", "POST"])
def index():

    if request.method == "POST":

        file = request.files["file"]

        groups, rooms, classes = parse_json(file)

        schedule = generate_schedule(groups, rooms, classes)

        excel_file = export_excel(schedule, groups)

        return send_file(excel_file, as_attachment=True)

    return render_template_string(HTML)


if __name__ == "__main__":
    app.run(debug=True)